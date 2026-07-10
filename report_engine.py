"""
=========================================================
Generator Raportu Nieobecności
Silnik programu
=========================================================

Autor: ChatGPT + Użytkownik
Wersja: 1.0
"""

from pathlib import Path
from openpyxl import load_workbook, Workbook
from datetime import date, datetime, timedelta
import calendar
import re


class ReportEngine:
    """
    Główna klasa programu.

    Odpowiada za:
        • wyszukanie wszystkich plików
        • analizę obecności
        • wygenerowanie raportu Excel
    """

    # =====================================================
    # Konstruktor klasy
    # =====================================================

    def __init__(self, folder, progress_callback=None, status_callback=None):

        # Folder wybrany przez użytkownika
        self.folder = Path(folder)

        # Funkcje aktualizacji GUI
        self.progress_callback = progress_callback
        self.status_callback = status_callback

        # Lista znalezionych plików
        self.files = []

        # Wyniki końcowe
        self.results = []

        # Miesiąc raportu
        self.report_month = None

        # Rok raportu
        self.report_year = None

    # =====================================================
    # Aktualizacja statusu
    # =====================================================

    def set_status(self, text):

        """
        Aktualizuje napis widoczny w oknie programu.
        """

        if self.status_callback:
            self.status_callback(text)

    # =====================================================
    # Aktualizacja paska postępu
    # =====================================================

    def set_progress(self, value):

        """
        Aktualizuje pasek postępu.
        """

        if self.progress_callback:
            self.progress_callback(value)

    # =====================================================
    # Wyszukiwanie plików Excel
    # =====================================================

    def scan_files(self):

        """
        Wyszukuje wszystkie pliki .xlsx
        również w podfolderach.

        Przykład:

        Czerwiec 2026
            |
            |-- Kowalski.xlsx
            |
            |-- Ukraina
                    |
                    |-- Iwan.xlsx

        Zostaną znalezione oba pliki.
        """

        self.files.clear()

        # rglob() przeszukuje wszystkie podfoldery
        for file in self.folder.rglob("*.xlsx"):

            # Pomijamy wygenerowane wcześniej raporty
            if file.name.lower().startswith("raport_nieobecności"):
                continue

            # Pomijamy pliki tymczasowe Excela
            if file.name.startswith("~$"):
                continue

            self.files.append(file)

        # Sortujemy alfabetycznie
        self.files.sort()

        return self.files

    # =====================================================
    # Wyciąganie nazwiska i imienia z nazwy pliku
    # =====================================================

    @staticmethod
    def get_employee_name(file_path):

        """
        Z:

            BORÓWKA PIOTR 168.xlsx

        otrzymujemy:

            BORÓWKA PIOTR

        Z:

            NOWAK JAN 232,5 delegacja.xlsx

        otrzymujemy:

            NOWAK JAN
        """

        name = Path(file_path).stem

        match = re.search(r"\s+\d", name)

        if match:
            return name[:match.start()].strip()

        return name.strip()

# =====================================================
    # Odczyt obecności z pliku Excel
    # =====================================================

    def read_attendance(self, excel_file):
        """
        Odczytuje wszystkie dni obecności pracownika.

        Zwraca:
            attendance_days -> zbiór dni (date)

        Przykład:

            2026-06-01
            2026-06-02
            2026-06-15

        Jeżeli zmiana trwała przez północ:

            30.06 22:00
            01.07 06:00

        zostaną dodane oba dni.
        """

        attendance_days = set()

        # Otwieramy plik tylko do odczytu.
        workbook = load_workbook(
            excel_file,
            read_only=True,
            data_only=True
        )

        # Dane zawsze są w pierwszym arkuszu.
        sheet = workbook.worksheets[0]

        # Przechodzimy po wszystkich wierszach.
        for row in sheet.iter_rows(min_row=1):

            cell_in = row[1].value      # kolumna B
            cell_out = row[2].value     # kolumna C

            # Pomijamy wiersze bez daty wejścia.
            if not isinstance(cell_in, datetime):
                continue

            # Zapamiętujemy miesiąc tylko raz.
            if self.report_month is None:
                self.report_month = cell_in.month
                self.report_year = cell_in.year

            start_day = cell_in.date()

            attendance_days.add(start_day)

            # Jeżeli istnieje data wyjścia.
            if isinstance(cell_out, datetime):

                end_day = cell_out.date()

                # Zmiana zakończyła się następnego dnia.
                if end_day > start_day:

                    current_day = start_day + timedelta(days=1)

                    while current_day <= end_day:

                        attendance_days.add(current_day)

                        current_day += timedelta(days=1)

        workbook.close()

        return attendance_days

  # =====================================================
    # Analiza obecności pracownika
    # =====================================================

    def analyse_employee(self, attendance_days):
        """
        Analizuje obecność pracownika.

        Zwraca:
            absent_days     -> lista dni roboczych
            saturday_days   -> lista sobót
        """

        absent_days = []
        saturday_days = []

        # Pierwszy dzień miesiąca
        first_day = date(
            self.report_year,
            self.report_month,
            1
        )

        # Liczba dni w miesiącu
        days_in_month = calendar.monthrange(
            self.report_year,
            self.report_month
        )[1]

        # Przechodzimy przez wszystkie dni miesiąca
        for day in range(1, days_in_month + 1):

            current_day = date(
                self.report_year,
                self.report_month,
                day
            )

            weekday = current_day.weekday()

            # -----------------------------------------
            # Niedzielę pomijamy
            # Monday = 0
            # Saturday = 5
            # Sunday = 6
            # -----------------------------------------

            if weekday == 6:
                continue

            # -----------------------------------------
            # Jeżeli pracownik był obecny
            # przechodzimy dalej
            # -----------------------------------------

            if current_day in attendance_days:
                continue

            # -----------------------------------------
            # Sobota
            # -----------------------------------------

            if weekday == 5:

                saturday_days.append(day)

            # -----------------------------------------
            # Dzień roboczy
            # -----------------------------------------

            else:

                absent_days.append(day)

        return absent_days, saturday_days

# =====================================================
    # Tworzenie raportu Excel
    # =====================================================

    def create_report(self):
        """
        Tworzy końcowy raport Excel.

        Kolumny:
            A - Pracownik
            B - Nieobecności
            C - Nieprzepracowane soboty
        """

        workbook = Workbook()
        sheet = workbook.active

        sheet.title = "Raport"

        # -------------------------------
        # Nagłówki
        # -------------------------------

        sheet["A1"] = "Pracownik"
        sheet["B1"] = "Nieobecności"
        sheet["C1"] = "Nieprzepracowane soboty"

        row = 2

        # Sortujemy pracowników alfabetycznie
        self.results.sort(key=lambda x: x["employee"])

        for employee in self.results:

            # Nazwa pracownika
            sheet.cell(row=row, column=1).value = employee["employee"]

            # Nieobecności
            if employee["absent"]:
                sheet.cell(row=row, column=2).value = ", ".join(
                    str(day) for day in employee["absent"]
                )
            else:
                sheet.cell(row=row, column=2).value = "brak"

            # Soboty
            if employee["saturdays"]:
                sheet.cell(row=row, column=3).value = ", ".join(
                    str(day) for day in employee["saturdays"]
                )
            else:
                sheet.cell(row=row, column=3).value = "brak"

            row += 1

        # -------------------------------
        # Automatyczna szerokość kolumn
        # -------------------------------

        for column in sheet.columns:

            length = 0

            column_letter = column[0].column_letter

            for cell in column:

                if cell.value is not None:

                    length = max(
                        length,
                        len(str(cell.value))
                    )

            sheet.column_dimensions[column_letter].width = length + 3

        # -------------------------------
        # Nazwa pliku
        # -------------------------------

        month_name = calendar.month_name[self.report_month]

        report_name = (
            f"Raport_nieobecności_"
            f"{month_name}_{self.report_year}.xlsx"
        )

        report_path = self.folder / report_name

        workbook.save(report_path)

        workbook.close()

        return report_path

    # =====================================================
    # Główna funkcja programu
    # =====================================================

    def run(self):
        """
        Główna funkcja programu.

        Kolejno:

            1. Wyszukuje pliki
            2. Analizuje każdy plik
            3. Tworzy raport
        """

        self.results.clear()

        self.set_status("Wyszukiwanie plików...")

        self.scan_files()

        total = len(self.files)

        if total == 0:
            raise Exception("Nie znaleziono żadnych plików Excel.")

        # -------------------------------------
        # Analiza wszystkich pracowników
        # -------------------------------------

        for index, excel_file in enumerate(self.files, start=1):

            employee = self.get_employee_name(excel_file)

            self.set_status(f"Analiza: {employee}")

            attendance = self.read_attendance(excel_file)

            absent, saturdays = self.analyse_employee(attendance)

            self.results.append({

                "employee": employee,

                "absent": absent,

                "saturdays": saturdays

            })

            progress = int(index / total * 100)

            self.set_progress(progress)

        # -------------------------------------
        # Tworzenie raportu
        # -------------------------------------

        self.set_status("Tworzenie raportu...")

        report = self.create_report()

        self.set_progress(100)

        self.set_status("Gotowe")

        return report


